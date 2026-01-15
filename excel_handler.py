# excel_handler.py - Excel import/export

import os
import asyncio
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ExcelHandler:
    """Excel fayllar bilan ishlash"""
    
    def __init__(self, db, excel_dir: str, export_dir: str):
        self.db = db
        self.excel_dir = excel_dir
        self.export_dir = export_dir
        os.makedirs(excel_dir, exist_ok=True)
        os.makedirs(export_dir, exist_ok=True)
    
    async def import_students(self, file_path: str) -> Dict[str, Any]:
        """
        Talabalar ma'lumotlarini import qilish
        "IT va ijtimoiy umumiy ro'yhad.xlsx" formati
        """
        result = {
            'success': False,
            'added': 0,
            'updated': 0,
            'errors': []
        }
        
        try:
            print(f"\n{'='*60}")
            print(f"IMPORT BOSHLANMOQDA: {os.path.basename(file_path)}")
            print(f"{'='*60}\n")
            
            # Excel faylni o'qish
            try:
                df = await asyncio.to_thread(
                    pd.read_excel,
                    file_path,
                    sheet_name=0,
                    header=None,
                    dtype=str,
                    engine='openpyxl'
                )
            except Exception as read_error:
                result['errors'].append(f"Fayl o'qishda xatolik: {str(read_error)}")
                return result
            
            print(f"✅ Fayl o'qildi: {len(df)} qator, {len(df.columns)} ustun\n")
            
            # Qator 2 dan boshlab (index 1) - birinchi qator sarlavha
            added = 0
            updated = 0
            
            for row_idx in range(1, len(df)):
                try:
                    row = df.iloc[row_idx]
                    
                    # Full Name (ustun 3 = index 2)
                    fullname = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else None
                    
                    # Agar ismi bo'sh yoki raqam bo'lsa o'tkazib yuborish
                    if not fullname or fullname.replace('.', '').isdigit() or fullname.lower() == 'nan':
                        continue
                    
                    # Passport raqamini tekshirish (ustun 11 = index 10)
                    passport = str(row.iloc[10]).strip().upper() if len(row) > 10 and pd.notna(row.iloc[10]) else None
                    if not passport or passport == 'NAN':
                        continue
                    
                    # JSHSHIR tekshirish (ustun 12 = index 11)
                    jshshir = str(row.iloc[11]).strip() if len(row) > 11 and pd.notna(row.iloc[11]) else None
                    if jshshir:
                        jshshir = jshshir.replace('.0', '').strip()
                        if jshshir.lower() == 'nan' or not jshshir.isdigit():
                            jshshir = None
                    
                    # Talaba ID (ustun 2 = index 1)
                    talaba_id = str(row.iloc[1]).strip() if len(row) > 1 and pd.notna(row.iloc[1]) else None
                    if talaba_id:
                        talaba_id = talaba_id.replace('.0', '').strip()
                        if talaba_id.lower() == 'nan':
                            talaba_id = None
                    
                    # Tug'ilgan sana (ustun 10 = index 9)
                    birth_date = None
                    if len(row) > 9 and pd.notna(row.iloc[9]):
                        bd = str(row.iloc[9]).strip()
                        if bd.lower() != 'nan':
                            birth_date = bd
                    
                    # Passport berilgan sana (ustun 13 = index 12)
                    passport_date = None
                    if len(row) > 12 and pd.notna(row.iloc[12]):
                        pd_date = str(row.iloc[12]).strip()
                        if pd_date.lower() != 'nan':
                            passport_date = pd_date
                    
                    # Kurs (ustun 14 = index 13)
                    course = None
                    if len(row) > 13 and pd.notna(row.iloc[13]):
                        course = str(row.iloc[13]).strip().replace('.0', '')
                        if course.lower() == 'nan':
                            course = None
                    
                    # Yordamchi funksiya
                    def safe_get(idx, default=None):
                        if len(row) > idx and pd.notna(row.iloc[idx]):
                            val = str(row.iloc[idx]).strip().replace('.0', '')
                            return val if val.lower() != 'nan' else default
                        return default
                    
                    student_data = {
                        'talaba_id': talaba_id,
                        'fullname': fullname,
                        'citizenship': safe_get(3),      # Fuqarolik (ustun 4)
                        'country': safe_get(4),          # Davlat (ustun 5)
                        'nationality': safe_get(5),      # Millat (ustun 6)
                        'region': safe_get(6),           # Viloyat (ustun 7)
                        'district': safe_get(7),         # Tuman (ustun 8)
                        'gender': safe_get(8),           # Jins (ustun 9)
                        'birth_date': birth_date,        # Tug'ilgan sana (ustun 10)
                        'passport': passport,            # Passport (ustun 11)
                        'jshshir': jshshir,              # JSHSHIR (ustun 12)
                        'passport_date': passport_date,  # Passport sanasi (ustun 13)
                        'course': course,                # Kurs (ustun 14)
                        'faculty': safe_get(14),         # Fakultet (ustun 15)
                        'group_name': safe_get(15),      # Guruh (ustun 16)
                        'language': safe_get(16),        # Ta'lim tili (ustun 17)
                        'study_year': safe_get(17),      # O'quv yili (ustun 18)
                        'semester': safe_get(18),        # Semestr (ustun 19)
                        'graduate': safe_get(19),        # Bitiruvchi (ustun 20)
                        'specialty': safe_get(20),       # Mutaxassislik (ustun 21)
                        'education_type': safe_get(21),  # Ta'lim turi (ustun 22)
                        'education_form': safe_get(22),  # Ta'lim shakli (ustun 23)
                        'payment_type': safe_get(23),    # To'lov turi (ustun 24)
                        'grant_type': safe_get(24),      # Grant turi (ustun 25)
                        'previous_education': safe_get(25),  # Oldingi ta'lim (ustun 26)
                        'student_category': safe_get(26),    # Talaba toifasi (ustun 27)
                        'social_category': safe_get(27),     # Ijtimoiy toifa (ustun 28)
                        'family_members': safe_get(28),      # Oila a'zolari (ustun 29)
                    }
                    
                    # Bazaga qo'shish
                    db_result = await self.db.add_student(student_data)
                    
                    if db_result['action'] == 'added':
                        added += 1
                    elif db_result['action'] == 'updated':
                        updated += 1
                    elif db_result['action'] == 'error':
                        result['errors'].append(f"Qator {row_idx + 1}: {db_result.get('error', 'Unknown')}")
                    
                except Exception as row_error:
                    result['errors'].append(f"Qator {row_idx + 1}: {str(row_error)}")
            
            result['success'] = True
            result['added'] = added
            result['updated'] = updated
            
            print(f"\n✅ IMPORT YAKUNLANDI")
            print(f"   Qo'shildi: {added}")
            print(f"   Yangilandi: {updated}")
            print(f"   Xatolar: {len(result['errors'])}")
            
        except Exception as e:
            result['errors'].append(f"Umumiy xatolik: {str(e)}")
            print(f"❌ XATOLIK: {str(e)}")
        
        return result
    
    async def export_survey_responses(self) -> Optional[str]:
        """So'rovnoma natijalarini Excel ga export qilish"""
        try:
            responses = await self.db.get_all_responses()
            
            if not responses:
                return None
            
            # Workbook yaratish
            wb = Workbook()
            ws = wb.active
            ws.title = "So'rovnoma natijalari"
            
            # Stil
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Headerlar
            headers = [
                "№",
                # Identifikatorlar
                "Unikal ID", "Talaba ID", "F.I.O",
                # Shaxsiy
                "Jinsi", "Tug'ilgan sana", "Passport", "JSHSHIR", "Fuqarolik",
                # Manzil
                "Viloyat", "Tuman",
                # O'qish
                "Kurs", "Fakultet", "Guruh", "Mutaxassislik",
                "Ta'lim turi", "Ta'lim shakli", "To'lov turi", "Grant turi",
                "Talaba toifasi", "Ijtimoiy toifa",
                # So'rovnoma javoblari
                "Telefon", "Doimiy manzil", "Doimiy joylashuv",
                "Oldingi ta'lim", "Hujjat raqami",
                "Yutuqlar bormi", "Yutuqlar",
                "Sertifikat bormi", "Sertifikat turi", "Sertifikat tafsiloti",
                "Grantga hujjat topshirganmi", "Grant tafsiloti",
                "Ijtimoiy himoya", "Temir daftar", "Yoshlar daftari",
                "Ota ismi", "Otasi hayotmi", "Ota telefoni",
                "Ona ismi", "Onasi hayotmi", "Ona telefoni", "Ota-onasi birga",
                "Yashash turi", "TTJ qayerdan", "Ijara manzili", "Ijara joylashuv", "Ijara egasi",
                "Ishlaydimi", "Ish joyi", "Oilalimi",
                "Xorijga chiqish pasporti", "Ijtimoiy tarmoq kanali", "Kanal/Guruh linklari",
                "So'rovnoma sanasi"
            ]
            
            # Header yozish
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Ma'lumotlar
            for row_idx, r in enumerate(responses, 2):
                data = [
                    row_idx - 1,
                    # Identifikatorlar
                    r.get('unique_id', ''),
                    r.get('talaba_id', ''),
                    r.get('fullname', ''),
                    # Shaxsiy
                    r.get('gender', ''),
                    r.get('birth_date', ''),
                    r.get('passport', ''),
                    r.get('jshshir', ''),
                    r.get('citizenship', ''),
                    # Manzil
                    r.get('region', ''),
                    r.get('district', ''),
                    # O'qish
                    r.get('course', ''),
                    r.get('faculty', ''),
                    r.get('group_name', ''),
                    r.get('specialty', ''),
                    r.get('education_type', ''),
                    r.get('education_form', ''),
                    r.get('payment_type', ''),
                    r.get('grant_type', ''),
                    r.get('student_category', ''),
                    r.get('social_category', ''),
                    # So'rovnoma javoblari
                    r.get('phone', ''),
                    r.get('permanent_address', ''),
                    r.get('permanent_location', ''),
                    r.get('previous_education', ''),
                    r.get('document_number', ''),
                    r.get('has_achievements', ''),
                    r.get('achievements', ''),
                    r.get('has_certificate', ''),
                    r.get('certificate_type', ''),
                    r.get('certificate_details', ''),
                    r.get('has_grant', ''),
                    r.get('grant_details', ''),
                    r.get('social_protection', ''),
                    r.get('iron_book', ''),
                    r.get('youth_book', ''),
                    r.get('father_name', ''),
                    r.get('father_alive', ''),
                    r.get('father_phone', ''),
                    r.get('mother_name', ''),
                    r.get('mother_alive', ''),
                    r.get('mother_phone', ''),
                    r.get('parents_together', ''),
                    r.get('living_type', ''),
                    r.get('ttj_location', ''),
                    r.get('rent_address', ''),
                    r.get('rent_location', ''),
                    r.get('rent_owner', ''),
                    r.get('is_working', ''),
                    r.get('workplace', ''),
                    r.get('is_married', ''),
                    r.get('has_foreign_passport', ''),
                    r.get('has_social_channels', ''),
                    r.get('social_links', ''),
                    r.get('created_at', '')
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
            
            # Ustun kengliklarini sozlash
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 18
            
            # Birinchi ustun (№) kichik
            ws.column_dimensions['A'].width = 5
            # F.I.O kattaroq
            ws.column_dimensions['D'].width = 30
            
            # Saqlash
            filename = f"sorovnoma_natijalari_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(self.export_dir, filename)
            
            await asyncio.to_thread(wb.save, filepath)
            
            return filepath
            
        except Exception as e:
            print(f"Export xatolik: {e}")
            return None
    
    async def export_students(self) -> Optional[str]:
        """Barcha talabalarni Excel ga export qilish (unikal ID bilan)"""
        try:
            students = await self.db.get_all_students()
            
            if not students:
                return None
            
            # Workbook yaratish
            wb = Workbook()
            ws = wb.active
            ws.title = "Talabalar"
            
            # Stil
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="217346", end_color="217346", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Headerlar
            headers = [
                "№", "Unikal ID", "Talaba ID", "F.I.O",
                "Fuqarolik", "Davlat", "Millat",
                "Viloyat", "Tuman", "Jinsi", "Tug'ilgan sana",
                "Passport", "JSHSHIR", "Passport sanasi",
                "Kurs", "Fakultet", "Guruh", "Ta'lim tili",
                "O'quv yili", "Semestr", "Bitiruvchi",
                "Mutaxassislik", "Ta'lim turi", "Ta'lim shakli",
                "To'lov turi", "Grant turi", "Oldingi ta'lim",
                "Talaba toifasi", "Ijtimoiy toifa", "Oila a'zolari", "Telefon",
                "Qo'shilgan vaqt", "Yangilangan vaqt"
            ]
            
            # Header yozish
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Ma'lumotlar
            for row_idx, s in enumerate(students, 2):
                data = [
                    row_idx - 1,
                    s.get('unique_id', ''),
                    s.get('talaba_id', ''),
                    s.get('fullname', ''),
                    s.get('citizenship', ''),
                    s.get('country', ''),
                    s.get('nationality', ''),
                    s.get('region', ''),
                    s.get('district', ''),
                    s.get('gender', ''),
                    s.get('birth_date', ''),
                    s.get('passport', ''),
                    s.get('jshshir', ''),
                    s.get('passport_date', ''),
                    s.get('course', ''),
                    s.get('faculty', ''),
                    s.get('group_name', ''),
                    s.get('language', ''),
                    s.get('study_year', ''),
                    s.get('semester', ''),
                    s.get('graduate', ''),
                    s.get('specialty', ''),
                    s.get('education_type', ''),
                    s.get('education_form', ''),
                    s.get('payment_type', ''),
                    s.get('grant_type', ''),
                    s.get('previous_education', ''),
                    s.get('student_category', ''),
                    s.get('social_category', ''),
                    s.get('family_members', ''),
                    s.get('phone', ''),
                    s.get('created_at', ''),
                    s.get('updated_at', ''),
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
            
            # Ustun kengliklarini sozlash
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 15
            
            # Birinchi ustun (№) kichik
            ws.column_dimensions['A'].width = 5
            # F.I.O kattaroq
            ws.column_dimensions['D'].width = 30
            
            # Saqlash
            filename = f"talabalar_royxati_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(self.export_dir, filename)
            
            await asyncio.to_thread(wb.save, filepath)
            
            return filepath
            
        except Exception as e:
            print(f"Export xatolik: {e}")
            return None
